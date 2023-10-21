''' The Indian Academy of Sciences announces their Summer Intership program every year and gives a list of projects available
    This code takes this data from the website and compiles it into an excel sheet for easier accessibiltiy. I am not sure if the
     website link still works '''

import requests
from bs4 import BeautifulSoup
import openpyxl as op

#set url and get webpage
url = "https://webjapps.ias.ac.in/fellowship2023/selectfellows.jsp?atype=B&area=%25"
req = requests.get(url)

#Get tags in webpage
page = BeautifulSoup(req.text, "html.parser")

#Main List has all the stuff in each cell of the table
main_stuff = list()

#These lists will segregate them based on subject
bio_stuff = list()
earth_stuff = list()
eng_stuff = list()
math_stuff = list()
chem_stuff = list()
phys_stuff = list()

#Some entries dont list any subject we store them in misc
misc_stuff = list()

#adding the cell contents to main stuff
for tag in page.find_all("td"):
    main_stuff.append(tag.text.strip())

#need to remove excess crap in the beginning
for i in range(15):
    main_stuff.pop(0)

#get total items in main_stuff
n = len(main_stuff)

#this will segregate the cells of the table into their corresponding rows
main_stuff_new_list = list()
temp_list = list()

for i in range(n):
    temp_list.append(main_stuff[i])
    if main_stuff[i] == "Accommodation Available" or main_stuff[i] == "Accommodation Not Available" :
        main_stuff_new_list.append(temp_list)
        temp_list = list()

#This goes through each row and segregates them based on subject
for i in main_stuff_new_list:
    i[1] = i[1].replace("/xa0"," ")
    i[1] = i[1].replace("/r/n"," ")
    #i = " ".join(i)
    #i = i.split()
    if "Chemistry" in i[3].split():
        chem_stuff.append(i)
    elif "Engineering" in i[3].split():
        eng_stuff.append(i)
    elif "Life" in i[3].split():
        bio_stuff.append(i)
    elif "Mathematics" in i[3].split():
        math_stuff.append(i)
    elif "Physics" in i[3].split():
        phys_stuff.append(i)
    elif "Earth" in i[3].split():
        earth_stuff.append(i)
    else:
        misc_stuff.append(i)

#forming excel sheet so I dont go insane
workbook = op.Workbook()

#Adding Biology Professors to the sheet
bio_sheet = workbook.create_sheet("Biology")
bio_sheet.title = "Biology Professors"

row_set = 1
for entry_list in bio_stuff:
    no_of_columns_needed = len(entry_list)
    for j in range(1,no_of_columns_needed+1):
        bio_sheet.cell(row=row_set, column = j).value = entry_list[j-1]
    print(str(row_set) + " Biology done")
    row_set = row_set + 1

print("Biology Done")

#Adding Earth sciences Professors
earth_sheet = workbook.create_sheet("Earth and Planetary Sciences")
earth_sheet.title = "ECS Professors"

row_set = 1
for entry_list in earth_stuff:
    no_of_columns_needed = len(entry_list)
    for j in range(1,no_of_columns_needed+1):
        earth_sheet.cell(row=row_set, column = j).value = entry_list[j-1]
    print(str(row_set) + " Earth done")
    row_set = row_set + 1

print("Earth Sciences Done")

#Adding Engineering Professors
eng_sheet = workbook.create_sheet("Engineering")
eng_sheet.title = "Engineering Professors"

row_set = 1
for entry_list in eng_stuff:
    no_of_columns_needed = len(entry_list)
    for j in range(1,no_of_columns_needed+1):
        eng_sheet.cell(row=row_set, column = j).value = entry_list[j-1]
    print(str(row_set) + " Engineering done")
    row_set = row_set + 1

print("Engineering Done")

#Adding Mathematics Professors
math_sheet = workbook.create_sheet("Mathematics")
math_sheet.title = "Mathematics Professors"
g
row_set = 1
for entry_list in math_stuff:
    no_of_columns_needed = len(entry_list)
    for j in range(1,no_of_columns_needed+1):
        math_sheet.cell(row=row_set, column = j).value = entry_list[j-1]
    print(str(row_set) + " Math done")
    row_set = row_set + 1

print("Math Done")

#Addings Chemistry Professors
chem_sheet = workbook.create_sheet("Chemistry")
chem_sheet.title = "Chemistry Professors"

row_set = 1
for entry_list in chem_stuff:
    no_of_columns_needed = len(entry_list)
    for j in range(1,no_of_columns_needed+1):
        chem_sheet.cell(row=row_set, column = j).value = entry_list[j-1]
    print(str(row_set) + " Chemistry done")
    row_set = row_set + 1g

print("Chemistry Done")

#Adding Physics Professors
phys_sheet = workbook.create_sheet("Physics")
phys_sheet.title = "Physics Professors"

row_set = 1
for entry_list in phys_stuff:
    no_of_columns_needed = len(entry_list)
    for j in range(1,no_of_columns_needed+1):
        phys_sheet.cell(row=row_set, column = j).value = entry_list[j-1]
    print(str(row_set) + " Physics done")
    row_set = row_set + 1

print("Physics Done")

#Adding Misc Professors
misc_sheet = workbook.create_sheet("Misc")
misc_sheet.title = "Professors That didn't list subject"

row_set = 1
for entry_list in misc_stuff:
    no_of_columns_needed = len(entry_list)
    for j in range(1,no_of_columns_needed+1):
        misc_sheet.cell(row=row_set, column = j).value = entry_list[j-1]
    print(str(row_set) + " Misc done")
    row_set = row_set + 1

print("Miscellaneous Professors Done")

#Saving the file
workbook.save(filename="C:\\Users\\saive\\Desktop\\fellows_sheet_better.xlsx")

#final checks to see amount of professors
print("The amount of Biology professors are " + str(len(bio_stuff)))
print("The amount of Physics professors are " + str(len(phys_stuff)))
print("The amount of ECS professors are " + str(len(earth_stuff)))
print("The amount of Math professors are " + str(len(math_stuff)))
print("The amount of Engineering professors are " + str(len(eng_stuff)))
print("The amount of Chemistry professors are " + str(len(chem_stuff)))
print("The amount of Inter-disciplinary professors are " + str(len(misc_stuff)))
total = len(misc_stuff)+len(bio_stuff) + len(earth_stuff) + len(eng_stuff) + len(math_stuff) + len(chem_stuff) + len(phys_stuff)
print("The total number of professors are "+str(total))
